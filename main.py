import tkinter as tk
import pandas as pd
import numpy as np
from tkinter.filedialog import asksaveasfilename
from tkinter.filedialog import askopenfilename
from openpyxl.workbook import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# rawdata = pd.read_csv(
#     r"D:\Github\Projects\Belmont-OCPD\Data\2019\csv\Class of 2019 First Destination Survey Raw Data no identifiers scrubbed version 3.3.csv",
#     encoding="ISO-8859-1")
# global cip
# cip = pd.read_csv(
#     r"D:\Github\Projects\Belmont-OCPD\Data\2019\CIP codes\Argos Report.csv", encoding="ISO-8859-1"
# )
# global institution
# institution = 'Belmont University'


def load_data_handshake():  # Loads the handshake data
    filename = askopenfilename()
    global rawdata
    rawdata = pd.read_csv(filename, encoding="ISO-8859-1")
    canvas1.create_window(100, 50, window=cip_button)
    button_import.destroy()


def load_cip_data():  # Loads the CIP data (typically from argos report)
    filename = askopenfilename()
    global cip
    cip = pd.read_csv(filename, encoding="ISO-8859-1")
    canvas1.create_window(100, 25, window=input_text)
    canvas1.create_window(100, 100, window=compile_document_button)
    canvas1.create_window(100, 50, window=institution_entry)
    cip_button.destroy()


def program_data_variables(major, education, rawdata, cip, institution):
    program_data_variables.institution = institution

    try:
        program_data_variables.CIP_code = cip.loc[cip.MajorDesc == major, 'CIPCode'].values[0]
    except:
        program_data_variables.CIP_code = np.nan

    program_data_variables.total_graduated = rawdata[(rawdata['Recipient Primary Major'] == major) &
                                                     (rawdata['Recipient Education Level'] == education)].shape[0]

    program_data_variables.full_time = rawdata[(rawdata['Employment Type'] == 'Full-Time') &
                                               (rawdata['Recipient Primary Major'] == major) &
                                               (rawdata['Recipient Education Level'] == education)].shape[0]

    program_data_variables.part_time = rawdata[(rawdata['Employment Type'] == 'Part-Time') &
                                               (rawdata['Recipient Primary Major'] == major) &
                                               (rawdata['Recipient Education Level'] == education)].shape[0]

    program_data_variables.entrepreneur_ft = rawdata[(rawdata['Employment Type'] == 'Full-Time') &
                                                     (rawdata['Recipient Primary Major'] == major) &
                                                     (rawdata['Recipient Education Level'] == education) &
                                                     (rawdata['Employment Category'] == 'Entrepreneur')].shape[0]

    program_data_variables.entrepreneur_pt = rawdata[(rawdata['Employment Type'] == 'Part-Time') &
                                                     (rawdata['Recipient Primary Major'] == major) &
                                                     (rawdata['Recipient Education Level'] == education) &
                                                     (rawdata['Employment Category'] == 'Entrepreneur')].shape[0]

    program_data_variables.temp_contract_ft = rawdata[(rawdata['Employment Type'] == 'Full-Time') &
                                                      (rawdata['Recipient Primary Major'] == major) &
                                                      (rawdata['Recipient Education Level'] == education) &
                                                      (rawdata['Employment Category'] == 'Temporary/Contract Work Assignment')].shape[0]

    program_data_variables.temp_contract_pt = rawdata[(rawdata['Employment Type'] == 'Part-Time') &
                                                      (rawdata['Recipient Primary Major'] == major) &
                                                      (rawdata['Recipient Education Level'] == education) &
                                                      (rawdata['Employment Category'] == 'Temporary/Contract Work Assignment')].shape[0]

    program_data_variables.freelance_ft = rawdata[(rawdata['Employment Type'] == 'Full-Time') &
                                                  (rawdata['Recipient Primary Major'] == major) &
                                                  (rawdata['Recipient Education Level'] == education) &
                                                  (rawdata['Employment Category'] == 'Freelancer')].shape[0]

    program_data_variables.freelance_pt = rawdata[(rawdata['Employment Type'] == 'Part-Time') &
                                                  (rawdata['Recipient Primary Major'] == major) &
                                                  (rawdata['Recipient Education Level'] == education) &
                                                  (rawdata['Employment Category'] == 'Freelancer')].shape[0]

    program_data_variables.faculty_tenure = rawdata[(rawdata['Recipient Education Level'] == education) &
                                                    (rawdata['Recipient Primary Major'] == major) &
                                                    (rawdata['Employment Category'] == 'Faculty Tenure')].shape[0]

    program_data_variables.faculty_nontenure = rawdata[(rawdata['Recipient Education Level'] == education) &
                                                       (rawdata['Recipient Primary Major'] == major) &
                                                       (rawdata['Employment Category'] == 'Faculty Non-Tenure')].shape[0]

    data1 = rawdata[(rawdata['Employment Type'] == 'Full-Time') &
                    (rawdata['Recipient Primary Major'] == major) &
                    (rawdata['Recipient Education Level'] == education) &
                    (rawdata['Internship'] == 'Yes')].shape[0]
    data2 = rawdata[(rawdata['Employment Type'] == 'Full-Time') &
                    (rawdata['Recipient Primary Major'] == major) &
                    (rawdata['Recipient Education Level'] == education) &
                    (rawdata['Is Fellowship?'] == 'Yes')].shape[0]
    program_data_variables.fellowship_intern_ft = data1 + data2

    data1 = rawdata[(rawdata['Employment Type'] == 'Part-Time') &
                    (rawdata['Recipient Primary Major'] == major) &
                    (rawdata['Recipient Education Level'] == education) &
                    (rawdata['Internship'] == 'Yes')].shape[0]
    data2 = rawdata[(rawdata['Employment Type'] == 'Part-Time') &
                    (rawdata['Recipient Primary Major'] == major) &
                    (rawdata['Recipient Education Level'] == education) &
                    (rawdata['Is Fellowship?'] == 'Yes')].shape[0]
    program_data_variables.fellowship_intern_pt = data1 + data2

    program_data_variables.outcome_service = rawdata[(rawdata['Recipient Primary Major'] == major) &
                                                     (rawdata['Recipient Education Level'] == education) &
                                                     (rawdata['Outcome'] == 'Volunteering')].shape[0]

    program_data_variables.outcome_military = rawdata[(rawdata['Recipient Primary Major'] == major) &
                                                      (rawdata['Recipient Education Level'] == education) &
                                                      (rawdata['Outcome'] == 'Military')].shape[0]

    program_data_variables.outcome_continue_education = rawdata[(rawdata['Recipient Primary Major'] == major) &
                                                                (rawdata['Recipient Education Level'] == education) &
                                                                (rawdata['Outcome'] == 'Continuing Education')].shape[0]

    program_data_variables.outcome_still_looking = rawdata[(rawdata['Recipient Primary Major'] == major) &
                                                           (rawdata['Recipient Education Level'] == education) &
                                                           (rawdata['Still Looking Option'] == 'Employment')].shape[0]

    program_data_variables.outcome_seeking_education = rawdata[(rawdata['Recipient Primary Major'] == major) &
                                                               (rawdata['Recipient Education Level'] == education) &
                                                               (rawdata['Still Looking Option'] == 'Continuing Education')].shape[0]

    program_data_variables.outcome_not_seeking = rawdata[(rawdata['Recipient Primary Major'] == major) &
                                                         (rawdata['Recipient Education Level'] == education) &
                                                         (rawdata['Outcome'] == 'Not Seeking')].shape[0]

    program_data_variables.no_info = rawdata[(rawdata['Recipient Primary Major'] == major) &
                                             (rawdata['Recipient Education Level'] == education) &
                                             (pd.isna(rawdata['Outcome']))].shape[0]

    program_data_variables.salaries_ft = rawdata[(rawdata['Recipient Primary Major'] == major) &
                                                 (rawdata['Employment Type'] == 'Full-Time') &
                                                 (rawdata['Recipient Education Level'] == education) &
                                                 (rawdata['Annual Salary'])].shape[0]

    data1 = rawdata[(rawdata['Recipient Primary Major'] == major) &
                    (rawdata['Employment Type'] == 'Full-Time') &
                    (rawdata['Recipient Education Level'] == education) &
                    (rawdata['Annual Salary'])]
    program_data_variables.salaries_mean = np.mean(data1['Annual Salary'])

    data1 = rawdata[(rawdata['Recipient Primary Major'] == major) &
                    (rawdata['Employment Type'] == 'Full-Time') &
                    (rawdata['Recipient Education Level'] == education) &
                    (rawdata['Annual Salary'])]
    program_data_variables.salaries_median = np.median(data1['Annual Salary'])

    data1 = rawdata[(rawdata['Recipient Primary Major'] == major) &
                    (rawdata['Employment Type'] == 'Full-Time') &
                    (rawdata['Recipient Education Level'] == education) &
                    (rawdata['Annual Salary'])]
    program_data_variables.salaries_low = (data1['Annual Salary'].min())

    data1 = rawdata[(rawdata['Recipient Primary Major'] == major) &
                    (rawdata['Employment Type'] == 'Full-Time') &
                    (rawdata['Recipient Education Level'] == education) &
                    (rawdata['Annual Salary'])]
    program_data_variables.salaries_high = (data1['Annual Salary'].max())

    program_data_variables.bonus = rawdata[(rawdata['Recipient Primary Major'] == major) &
                                           (rawdata['Employment Type'] == 'Full-Time') &
                                           (rawdata['Recipient Education Level'] == education) &
                                           (rawdata['Bonus Amount'] != 0) & (pd.notna(rawdata['Bonus Amount']))].shape[0]

    data1 = rawdata[(rawdata['Recipient Primary Major'] == major) &
                    (rawdata['Employment Type'] == 'Full-Time') &
                    (rawdata['Recipient Education Level'] == education) &
                    (rawdata['Bonus Amount'] != 0) & (pd.notna(rawdata['Bonus Amount']))]
    program_data_variables.bonus_mean = np.mean(data1['Bonus Amount'])

    data1 = rawdata[(rawdata['Recipient Primary Major'] == major) &
                    (rawdata['Employment Type'] == 'Full-Time') &
                    (rawdata['Recipient Education Level'] == education) &
                    (rawdata['Bonus Amount'] != 0) & (pd.notna(rawdata['Bonus Amount']))]
    program_data_variables.bonus_median = np.median(data1['Bonus Amount'])

    data1 = rawdata[(rawdata['Recipient Primary Major'] == major) &
                    (rawdata['Employment Type'] == 'Full-Time') &
                    (rawdata['Recipient Education Level'] == education) &
                    (rawdata['Bonus Amount'] != 0) & (pd.notna(rawdata['Bonus Amount']))]
    program_data_variables.bonus_low = data1['Bonus Amount'].min()

    data1 = rawdata[(rawdata['Recipient Primary Major'] == major) &
                    (rawdata['Employment Type'] == 'Full-Time') &
                    (rawdata['Recipient Education Level'] == education) &
                    (rawdata['Bonus Amount'] != 0) & (pd.notna(rawdata['Bonus Amount']))]
    program_data_variables.bonus_high = data1['Bonus Amount'].max()
    return ()


def overall_data_variables(education, rawdata, institution):
    overall_data_variables.institution = institution

    overall_data_variables.total_graduated = rawdata[(rawdata['Recipient Education Level'] == education)].shape[0]

    overall_data_variables.ft = rawdata[(rawdata['Employment Type'] == 'Full-Time') &
                                        (rawdata['Recipient Education Level'] == education)].shape[0]

    overall_data_variables.pt = rawdata[(rawdata['Employment Type'] == 'Part-Time') &
                                        (rawdata['Recipient Education Level'] == education)].shape[0]

    overall_data_variables.faculty_tenure = rawdata[(rawdata['Recipient Education Level'] == education) &
                                                    (rawdata['Employment Category'] == 'Faculty Tenure')].shape[0]

    overall_data_variables.faculty_nontenure = rawdata[(rawdata['Recipient Education Level'] == education) &
                                                       (rawdata['Employment Category'] == 'Faculty Non-Tenure')].shape[0]

    overall_data_variables.entrepreneur_ft = rawdata[(rawdata['Employment Type'] == 'Full-Time') &
                                                     (rawdata['Recipient Education Level'] == education) &
                                                     (rawdata['Employment Category'] == 'Entrepreneur')].shape[0]

    overall_data_variables.entrepreneur_pt = rawdata[(rawdata['Employment Type'] == 'Part-Time') &
                                                     (rawdata['Recipient Education Level'] == education) &
                                                     (rawdata['Employment Category'] == 'Entrepreneur')].shape[0]

    overall_data_variables.temp_contract_ft = rawdata[(rawdata['Employment Type'] == 'Full-Time') &
                                                      (rawdata['Recipient Education Level'] == education) &
                                                      (rawdata['Employment Category'] == 'Temporary/Contract Work Assignment')].shape[0]

    overall_data_variables.temp_contract_pt = rawdata[(rawdata['Employment Type'] == 'Part-Time') &
                                                      (rawdata['Recipient Education Level'] == education) &
                                                      (rawdata['Employment Category'] == 'Temporary/Contract Work Assignment')].shape[0]

    overall_data_variables.freelance_ft = rawdata[(rawdata['Employment Type'] == 'Full-Time') &
                                                  (rawdata['Recipient Education Level'] == education) &
                                                  (rawdata['Employment Category'] == 'Freelancer')].shape[0]

    overall_data_variables.freelance_pt = rawdata[(rawdata['Employment Type'] == 'Part-Time') &
                                                  (rawdata['Recipient Education Level'] == education) &
                                                  (rawdata['Employment Category'] == 'Freelancer')].shape[0]

    data1 = rawdata[(rawdata['Employment Type'] == 'Full-Time') &
                    (rawdata['Recipient Education Level'] == education) &
                    (rawdata['Internship'] == 'Yes')].shape[0]
    data2 = rawdata[(rawdata['Employment Type'] == 'Full-Time') &
                    (rawdata['Recipient Education Level'] == education) &
                    (rawdata['Is Fellowship?'] == 'Yes')].shape[0]
    overall_data_variables.fellowship_intern_ft = data1 + data2

    data1 = rawdata[(rawdata['Employment Type'] == 'Part-Time') &
                    (rawdata['Recipient Education Level'] == education) &
                    (rawdata['Internship'] == 'Yes')].shape[0]
    data2 = rawdata[(rawdata['Employment Type'] == 'Part-Time') &
                    (rawdata['Recipient Education Level'] == education) &
                    (rawdata['Is Fellowship?'] == 'Yes')].shape[0]
    overall_data_variables.fellowship_intern_pt = data1 + data2

    overall_data_variables.outcome_service = rawdata[(rawdata['Recipient Education Level'] == education) &
                                                     (rawdata['Outcome'] == 'Volunteering')].shape[0]

    overall_data_variables.outcome_military = rawdata[(rawdata['Recipient Education Level'] == education) &
                                                      (rawdata['Outcome'] == 'Military')].shape[0]

    overall_data_variables.outcome_continue_education = rawdata[(rawdata['Recipient Education Level'] == education) &
                                                                (rawdata['Outcome'] == 'Continuing Education')].shape[0]

    overall_data_variables.outcome_still_looking = rawdata[(rawdata['Recipient Education Level'] == education) &
                                                           (rawdata['Still Looking Option'] == 'Employment')].shape[0]

    overall_data_variables.outcome_seeking_education = rawdata[(rawdata['Recipient Education Level'] == education) &
                                                               (rawdata['Still Looking Option'] == 'Continuing Education')].shape[0]

    overall_data_variables.outcome_not_seeking = rawdata[(rawdata['Recipient Education Level'] == education) &
                                                         (rawdata['Outcome'] == 'Not Seeking')].shape[0]

    overall_data_variables.no_info = rawdata[(rawdata['Recipient Education Level'] == education) &
                                             (pd.isna(rawdata['Outcome']))].shape[0]

    overall_data_variables.salaries_ft = rawdata[(rawdata['Recipient Education Level'] == education) &
                                                 (rawdata['Employment Type'] == 'Full-Time') &
                                                 (rawdata['Annual Salary'])].shape[0]

    data = rawdata[(rawdata['Recipient Education Level'] == education) &
                   (rawdata['Employment Type'] == 'Full-Time') &
                   (rawdata['Annual Salary'])]
    overall_data_variables.salaries_mean = np.mean(data['Annual Salary'])

    data = rawdata[(rawdata['Recipient Education Level'] == education) &
                   (rawdata['Employment Type'] == 'Full-Time') &
                   (rawdata['Annual Salary'])]
    overall_data_variables.salaries_median = np.median(data['Annual Salary'])

    overall_data_variables.bonus = rawdata[(rawdata['Recipient Education Level'] == education) &
                                           (rawdata['Employment Type'] == 'Full-Time') &
                                           (rawdata['Bonus Amount'] != 0) & (pd.notna(rawdata['Bonus Amount']))].shape[0]

    data = rawdata[(rawdata['Recipient Education Level'] == education) &
                   (rawdata['Employment Type'] == 'Full-Time') &
                   (rawdata['Bonus Amount'] != 0) &
                   (pd.notna(rawdata['Bonus Amount']))]
    overall_data_variables.bonus_mean = np.mean(data['Bonus Amount'])

    data = rawdata[(rawdata['Recipient Education Level'] == education) &
                   (rawdata['Employment Type'] == 'Full-Time') &
                   (rawdata['Bonus Amount'] != 0) &
                   (pd.notna(rawdata['Bonus Amount']))]
    overall_data_variables.bonus_median = np.median(data['Bonus Amount'])
    return ()


def associate_program_data_function():
    education = 'Associate'
    education_level = rawdata[rawdata["Recipient Education Level"].str.startswith(education, na=False)]
    majors = education_level['Recipient Primary Major'].unique()
    majors = pd.DataFrame(majors, columns=['major'])
    df = pd.DataFrame(columns=['Institution Name',
                               'Academic Program Name',
                               'CIP Code',
                               'Total Graduated',
                               'Full-Time',
                               'Part-Time',
                               'Entrepreneur Full-Time',
                               'Entrepreneur Part-Time',
                               'Temp/Contract FT',
                               'Temp/Contract PT',
                               'Freelance  FT',
                               'Freelance PT',
                               'Fellowship/Intern FT',
                               'Fellowship/Intern PT',
                               'service',
                               'Military',
                               'Continuing Education',
                               'Seeking Employment',
                               'Seeking Education',
                               'Not Seek',
                               'no info',
                               '# of Salaries (Full-time Employed)',
                               'Salary Mean',
                               'Median Salary',
                               'Salary Low',
                               'Salary High',
                               'Bonus Numbers',
                               'Bonus Mean',
                               'Bonus Median',
                               'Bonus Low',
                               'Bonus High'])
    for i in majors['major']:
        program_data_variables(i, education, rawdata=rawdata, cip=cip, institution=institution)
        data_list = []
        df_list = [program_data_variables.institution,
                   i,
                   program_data_variables.CIP_code,
                   program_data_variables.total_graduated,
                   program_data_variables.full_time,
                   program_data_variables.part_time,
                   program_data_variables.entrepreneur_ft,
                   program_data_variables.entrepreneur_pt,
                   program_data_variables.temp_contract_ft,
                   program_data_variables.temp_contract_pt,
                   program_data_variables.freelance_ft,
                   program_data_variables.freelance_pt,
                   program_data_variables.fellowship_intern_ft,
                   program_data_variables.fellowship_intern_pt,
                   program_data_variables.outcome_service,
                   program_data_variables.outcome_military,
                   program_data_variables.outcome_continue_education,
                   program_data_variables.outcome_still_looking,
                   program_data_variables.outcome_seeking_education,
                   program_data_variables.outcome_not_seeking,
                   program_data_variables.no_info,
                   program_data_variables.salaries_ft,
                   program_data_variables.salaries_mean,
                   program_data_variables.salaries_median,
                   program_data_variables.salaries_low,
                   program_data_variables.salaries_high,
                   program_data_variables.bonus,
                   program_data_variables.bonus_mean,
                   program_data_variables.bonus_median,
                   program_data_variables.bonus_low,
                   program_data_variables.bonus_high
                   ]
        for var in df_list:
            data_list.append(var)
        s = pd.Series(data_list, index=df.columns)
        df = df.append(s, ignore_index=True)
    return df


def associate_overall_data_function():
    df = pd.DataFrame(columns=['Institution Name',
                               'Total Graduated',
                               'Full-Time',
                               'Part-Time',
                               'Entrepreneur Full-Time',
                               'Entrepreneur Part-Time',
                               'Temp/Contract FT',
                               'Temp/Contract PT',
                               'Freelance  FT',
                               'Freelance PT',
                               'Fellowship/Intern FT',
                               'Fellowship/Intern PT',
                               'Service',
                               'Military',
                               'Continuing Education',
                               'Seeking Employment',
                               'Seeking Education',
                               'Not Seek',
                               'no info',
                               '# of Salaries (Full-time Employed)',
                               'Salary Mean',
                               'Median Salary',
                               'Receiving Bonus',
                               'Bonus Mean',
                               'Bonus Median'])
    data_list = []
    education = 'Associate'
    overall_data_variables(education, rawdata=rawdata, institution=institution)
    df_list = [overall_data_variables.institution,
               overall_data_variables.total_graduated,
               overall_data_variables.ft,
               overall_data_variables.pt,
               overall_data_variables.entrepreneur_ft,
               overall_data_variables.entrepreneur_pt,
               overall_data_variables.temp_contract_ft,
               overall_data_variables.temp_contract_pt,
               overall_data_variables.freelance_ft,
               overall_data_variables.freelance_pt,
               overall_data_variables.fellowship_intern_ft,
               overall_data_variables.fellowship_intern_pt,
               overall_data_variables.outcome_service,
               overall_data_variables.outcome_military,
               overall_data_variables.outcome_continue_education,
               overall_data_variables.outcome_still_looking,
               overall_data_variables.outcome_seeking_education,
               overall_data_variables.outcome_not_seeking,
               overall_data_variables.no_info,
               overall_data_variables.salaries_ft,
               overall_data_variables.salaries_mean,
               overall_data_variables.salaries_median,
               overall_data_variables.bonus,
               overall_data_variables.bonus_mean,
               overall_data_variables.bonus_median
               ]
    for i in df_list:
        data_list.append(i)
    s = pd.Series(data_list, index=df.columns)
    df = df.append(s, ignore_index=True)
    return df


def bachelors_program_data_function():
    education = 'Bachelors'
    education_level = rawdata[rawdata["Recipient Education Level"].str.startswith(education, na=False)]
    majors = education_level['Recipient Primary Major'].unique()
    majors = pd.DataFrame(majors, columns=['major'])
    df = pd.DataFrame(columns=['Institution Name',
                               'Academic Program Name',
                               'CIP Code',
                               'Total Graduated',
                               'Full-Time',
                               'Part-Time',
                               'Entrepreneur Full-Time',
                               'Entrepreneur Part-Time',
                               'Temp/Contract FT',
                               'Temp/Contract PT',
                               'Freelance  FT',
                               'Freelance PT',
                               'Fellowship/Intern FT',
                               'Fellowship/Intern PT',
                               'service',
                               'Military',
                               'Continuing Education',
                               'Seeking Employment',
                               'Seeking Education',
                               'Not Seek',
                               'no info',
                               '# of Salaries (Full-time Employed)',
                               'Salary Mean',
                               'Median Salary',
                               'Salary Low',
                               'Salary High',
                               'Bonus Numbers',
                               'Bonus Mean',
                               'Bonus Median',
                               'Bonus Low',
                               'Bonus High'])
    for i in majors['major']:
        program_data_variables(i, education, rawdata=rawdata, cip=cip, institution=institution)
        data_list = []
        df_list = [program_data_variables.institution,
                   i,
                   program_data_variables.CIP_code,
                   program_data_variables.total_graduated,
                   program_data_variables.full_time,
                   program_data_variables.part_time,
                   program_data_variables.entrepreneur_ft,
                   program_data_variables.entrepreneur_pt,
                   program_data_variables.temp_contract_ft,
                   program_data_variables.temp_contract_pt,
                   program_data_variables.freelance_ft,
                   program_data_variables.freelance_pt,
                   program_data_variables.fellowship_intern_ft,
                   program_data_variables.fellowship_intern_pt,
                   program_data_variables.outcome_service,
                   program_data_variables.outcome_military,
                   program_data_variables.outcome_continue_education,
                   program_data_variables.outcome_still_looking,
                   program_data_variables.outcome_seeking_education,
                   program_data_variables.outcome_not_seeking,
                   program_data_variables.no_info,
                   program_data_variables.salaries_ft,
                   program_data_variables.salaries_mean,
                   program_data_variables.salaries_median,
                   program_data_variables.salaries_low,
                   program_data_variables.salaries_high,
                   program_data_variables.bonus,
                   program_data_variables.bonus_mean,
                   program_data_variables.bonus_median,
                   program_data_variables.bonus_low,
                   program_data_variables.bonus_high
                   ]
        for var in df_list:
            data_list.append(var)
        s = pd.Series(data_list, index=df.columns)
        df = df.append(s, ignore_index=True)
    return df


def bachelors_overall_data_function():
    df = pd.DataFrame(columns=['Institution Name',
                               'Total Graduated',
                               'Full-Time',
                               'Part-Time',
                               'Entrepreneur Full-Time',
                               'Entrepreneur Part-Time',
                               'Temp/Contract FT',
                               'Temp/Contract PT',
                               'Freelance  FT',
                               'Freelance PT',
                               'Fellowship/Intern FT',
                               'Fellowship/Intern PT',
                               'Service',
                               'Military',
                               'Continuing Education',
                               'Seeking Employment',
                               'Seeking Education',
                               'Not Seek',
                               'no info',
                               '# of Salaries (Full-time Employed)',
                               'Salary Mean',
                               'Median Salary',
                               'Receiving Bonus',
                               'Bonus Mean',
                               'Bonus Median'])
    data_list = []
    education = 'Bachelors'
    overall_data_variables(education, rawdata=rawdata, institution=institution)
    df_list = [overall_data_variables.institution,
               overall_data_variables.total_graduated,
               overall_data_variables.ft,
               overall_data_variables.pt,
               overall_data_variables.entrepreneur_ft,
               overall_data_variables.entrepreneur_pt,
               overall_data_variables.temp_contract_ft,
               overall_data_variables.temp_contract_pt,
               overall_data_variables.freelance_ft,
               overall_data_variables.freelance_pt,
               overall_data_variables.fellowship_intern_ft,
               overall_data_variables.fellowship_intern_pt,
               overall_data_variables.outcome_service,
               overall_data_variables.outcome_military,
               overall_data_variables.outcome_continue_education,
               overall_data_variables.outcome_still_looking,
               overall_data_variables.outcome_seeking_education,
               overall_data_variables.outcome_not_seeking,
               overall_data_variables.no_info,
               overall_data_variables.salaries_ft,
               overall_data_variables.salaries_mean,
               overall_data_variables.salaries_median,
               overall_data_variables.bonus,
               overall_data_variables.bonus_mean,
               overall_data_variables.bonus_median
               ]
    for i in df_list:
        data_list.append(i)
    s = pd.Series(data_list, index=df.columns)
    df = df.append(s, ignore_index=True)
    return df


def masters_program_data_function():
    education = 'Masters'
    education_level = rawdata[rawdata["Recipient Education Level"].str.startswith(education, na=False)]
    majors = education_level['Recipient Primary Major'].unique()
    majors = pd.DataFrame(majors, columns=['major'])
    df = pd.DataFrame(columns=['Institution Name',
                               'Academic Program Name',
                               'CIP Code',
                               'Total Graduated',
                               'Full-Time',
                               'Part-Time',
                               'Faculty Tenure Track',
                               'Faculty Non-Tenure Track',
                               'Entrepreneur Full-Time',
                               'Entrepreneur Part-Time',
                               'Temp/Contract FT',
                               'Temp/Contract PT',
                               'Freelance  FT',
                               'Freelance PT',
                               'Fellowship/Intern FT',
                               'Fellowship/Intern PT',
                               'service',
                               'Military',
                               'Continuing Education',
                               'Seeking Employment',
                               'Seeking Education',
                               'Not Seek',
                               'no info',
                               '# of Salaries (Full-time Employed)',
                               'Salary Mean',
                               'Median Salary',
                               'Bonus Numbers',
                               'Bonus Mean',
                               'Bonus Median',
                               ])
    for i in majors['major']:
        program_data_variables(i, education, rawdata=rawdata, cip=cip, institution=institution)
        data_list = []
        df_list = [program_data_variables.institution,
                   i,
                   program_data_variables.CIP_code,
                   program_data_variables.total_graduated,
                   program_data_variables.full_time,
                   program_data_variables.part_time,
                   program_data_variables.faculty_tenure,
                   program_data_variables.faculty_nontenure,
                   program_data_variables.entrepreneur_ft,
                   program_data_variables.entrepreneur_pt,
                   program_data_variables.temp_contract_ft,
                   program_data_variables.temp_contract_pt,
                   program_data_variables.freelance_ft,
                   program_data_variables.freelance_pt,
                   program_data_variables.fellowship_intern_ft,
                   program_data_variables.fellowship_intern_pt,
                   program_data_variables.outcome_service,
                   program_data_variables.outcome_military,
                   program_data_variables.outcome_continue_education,
                   program_data_variables.outcome_still_looking,
                   program_data_variables.outcome_seeking_education,
                   program_data_variables.outcome_not_seeking,
                   program_data_variables.no_info,
                   program_data_variables.salaries_ft,
                   program_data_variables.salaries_mean,
                   program_data_variables.salaries_median,
                   program_data_variables.bonus,
                   program_data_variables.bonus_mean,
                   program_data_variables.bonus_median,
                   ]
        for var in df_list:
            data_list.append(var)
        s = pd.Series(data_list, index=df.columns)
        df = df.append(s, ignore_index=True)
    return df


def masters_overall_data_function():
    df = pd.DataFrame(columns=['Institution Name',
                               'Total Graduated',
                               'Full-Time',
                               'Part-Time',
                               '# Faculty Tenure Track',
                               '# Faculty Non-Tenure Track',
                               'Entrepreneur Full-Time',
                               'Entrepreneur Part-Time',
                               'Temp/Contract FT',
                               'Temp/Contract PT',
                               'Freelance  FT',
                               'Freelance PT',
                               'Fellowship/Intern FT',
                               'Fellowship/Intern PT',
                               'Service',
                               'Military',
                               'Continuing Education',
                               'Seeking Employment',
                               'Seeking Education',
                               'Not Seek',
                               'no info',
                               '# of Salaries (Full-time Employed)',
                               'Salary Mean',
                               'Median Salary',
                               'Receiving Bonus',
                               'Bonus Mean',
                               'Bonus Median'])
    data_list = []
    education = 'Masters'
    overall_data_variables(education, rawdata=rawdata, institution=institution)
    df_list = [overall_data_variables.institution,
               overall_data_variables.total_graduated,
               overall_data_variables.ft,
               overall_data_variables.pt,
               overall_data_variables.faculty_tenure,
               overall_data_variables.faculty_nontenure,
               overall_data_variables.entrepreneur_ft,
               overall_data_variables.entrepreneur_pt,
               overall_data_variables.temp_contract_ft,
               overall_data_variables.temp_contract_pt,
               overall_data_variables.freelance_ft,
               overall_data_variables.freelance_pt,
               overall_data_variables.fellowship_intern_ft,
               overall_data_variables.fellowship_intern_pt,
               overall_data_variables.outcome_service,
               overall_data_variables.outcome_military,
               overall_data_variables.outcome_continue_education,
               overall_data_variables.outcome_still_looking,
               overall_data_variables.outcome_seeking_education,
               overall_data_variables.outcome_not_seeking,
               overall_data_variables.no_info,
               overall_data_variables.salaries_ft,
               overall_data_variables.salaries_mean,
               overall_data_variables.salaries_median,
               overall_data_variables.bonus,
               overall_data_variables.bonus_mean,
               overall_data_variables.bonus_median
               ]
    for i in df_list:
        data_list.append(i)
    s = pd.Series(data_list, index=df.columns)
    df = df.append(s, ignore_index=True)
    return df


def doctorate_program_data_function():
    education = 'Doctorate'
    education_level = rawdata[rawdata["Recipient Education Level"].str.startswith(education, na=False)]
    majors = education_level['Recipient Primary Major'].unique()
    majors = pd.DataFrame(majors, columns=['major'])
    df = pd.DataFrame(columns=['Institution Name',
                               'Academic Program Name',
                               'CIP Code',
                               'Total Graduated',
                               'Full-Time',
                               'Part-Time',
                               'Faculty Tenure Track',
                               'Faculty Non-Tenure Track',
                               'Entrepreneur Full-Time',
                               'Entrepreneur Part-Time',
                               'Temp/Contract FT',
                               'Temp/Contract PT',
                               'Freelance  FT',
                               'Freelance PT',
                               'Fellowship/Intern FT',
                               'Fellowship/Intern PT',
                               'service',
                               'Military',
                               'Continuing Education',
                               'Seeking Employment',
                               'Seeking Education',
                               'Not Seek',
                               'no info',
                               '# of Salaries (Full-time Employed)',
                               'Salary Mean',
                               'Median Salary',
                               'Bonus Numbers',
                               'Bonus Mean',
                               'Bonus Median',
                               ])
    for i in majors['major']:
        program_data_variables(i, education, rawdata=rawdata, cip=cip, institution=institution)
        data_list = []
        df_list = [program_data_variables.institution,
                   i,
                   program_data_variables.CIP_code,
                   program_data_variables.total_graduated,
                   program_data_variables.full_time,
                   program_data_variables.part_time,
                   program_data_variables.faculty_tenure,
                   program_data_variables.faculty_nontenure,
                   program_data_variables.entrepreneur_ft,
                   program_data_variables.entrepreneur_pt,
                   program_data_variables.temp_contract_ft,
                   program_data_variables.temp_contract_pt,
                   program_data_variables.freelance_ft,
                   program_data_variables.freelance_pt,
                   program_data_variables.fellowship_intern_ft,
                   program_data_variables.fellowship_intern_pt,
                   program_data_variables.outcome_service,
                   program_data_variables.outcome_military,
                   program_data_variables.outcome_continue_education,
                   program_data_variables.outcome_still_looking,
                   program_data_variables.outcome_seeking_education,
                   program_data_variables.outcome_not_seeking,
                   program_data_variables.no_info,
                   program_data_variables.salaries_ft,
                   program_data_variables.salaries_mean,
                   program_data_variables.salaries_median,
                   program_data_variables.bonus,
                   program_data_variables.bonus_mean,
                   program_data_variables.bonus_median,
                   ]
        for var in df_list:
            data_list.append(var)
        s = pd.Series(data_list, index=df.columns)
        df = df.append(s, ignore_index=True)
    # filename = asksaveasfile(defaultextension=".csv",
    #                         filetypes=(("Comma-separated values file", "*.csv"),
    #                                   ("All Files", "*.*")))
    # df.to_csv(filename, line_terminator='\n')
    return df


def doctorate_overall_data_function():
    df = pd.DataFrame(columns=['Institution Name',
                               'Total Graduated',
                               'Full-Time',
                               'Part-Time',
                               '# Faculty Tenure Track',
                               '# Faculty Non-Tenure Track',
                               'Entrepreneur Full-Time',
                               'Entrepreneur Part-Time',
                               'Temp/Contract FT',
                               'Temp/Contract PT',
                               'Freelance  FT',
                               'Freelance PT',
                               'Fellowship/Intern FT',
                               'Fellowship/Intern PT',
                               'Service',
                               'Military',
                               'Continuing Education',
                               'Seeking Employment',
                               'Seeking Education',
                               'Not Seek',
                               'no info',
                               '# of Salaries (Full-time Employed)',
                               'Salary Mean',
                               'Median Salary',
                               'Receiving Bonus',
                               'Bonus Mean',
                               'Bonus Median'])
    data_list = []
    education = 'Doctorate'
    overall_data_variables(education, rawdata=rawdata, institution=institution)
    df_list = [overall_data_variables.institution,
               overall_data_variables.total_graduated,
               overall_data_variables.ft,
               overall_data_variables.pt,
               overall_data_variables.faculty_tenure,
               overall_data_variables.faculty_nontenure,
               overall_data_variables.entrepreneur_ft,
               overall_data_variables.entrepreneur_pt,
               overall_data_variables.temp_contract_ft,
               overall_data_variables.temp_contract_pt,
               overall_data_variables.freelance_ft,
               overall_data_variables.freelance_pt,
               overall_data_variables.fellowship_intern_ft,
               overall_data_variables.fellowship_intern_pt,
               overall_data_variables.outcome_service,
               overall_data_variables.outcome_military,
               overall_data_variables.outcome_continue_education,
               overall_data_variables.outcome_still_looking,
               overall_data_variables.outcome_seeking_education,
               overall_data_variables.outcome_not_seeking,
               overall_data_variables.no_info,
               overall_data_variables.salaries_ft,
               overall_data_variables.salaries_mean,
               overall_data_variables.salaries_median,
               overall_data_variables.bonus,
               overall_data_variables.bonus_mean,
               overall_data_variables.bonus_median
               ]
    for i in df_list:
        data_list.append(i)
    s = pd.Series(data_list, index=df.columns)
    df = df.append(s, ignore_index=True)
    return df


def compile_document():
    global institution
    institution = institution_entry.get()

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Associate's Summary"
    ws2 = wb.create_sheet()
    ws2.title = "Bachelor's Summary"
    ws3 = wb.create_sheet()
    ws3.title = "Master's Summary"
    ws4 = wb.create_sheet()
    ws4.title = "Ph.D. Summary"
    ws5 = wb.create_sheet()
    ws5.title = "Program Data Associate's"
    ws6 = wb.create_sheet()
    ws6.title = "Program Data - Bachelor's"
    ws7 = wb.create_sheet()
    ws7.title = "Program Data - Master's"
    ws8 = wb.create_sheet()
    ws8.title = "Program Data - Doctoral"
    ws_list = [[ws1, associate_overall_data_function()],
               [ws2, bachelors_overall_data_function()],
               [ws3, masters_overall_data_function()],
               [ws4, doctorate_overall_data_function()],
               [ws5, associate_program_data_function()],
               [ws6, bachelors_program_data_function()],
               [ws7, masters_program_data_function()],
               [ws8, doctorate_program_data_function()]
               ]
    for sheets in ws_list:
        try:
            for r in dataframe_to_rows(sheets[1], index=False, header=True):
                sheets[0].append(r)
        except:
            pass
    filename = asksaveasfilename(defaultextension=".xlsx",
                                 filetypes=(("Microsoft Excel Open XML Spreadsheet File", "*.xlsx"),
                                            ("All Files", "*.*")))
    wb.save(filename)


if __name__ == "__main__":
    root = tk.Tk()
    root.title("NACE Compiler")
    canvas1 = tk.Canvas(root, width=200, height=200)
    canvas1.pack()
    button_import = tk.Button(root, text='Select .CSV Handshake Data', command=load_data_handshake)
    canvas1.create_window(100, 50, window=button_import)
    input_text = tk.Label(root, text='Input Institution name:')

    cip_button = tk.Button(root, text='Select .CSV CIP Data (argos report)', command=load_cip_data)
    institution_entry = tk.Entry(root)  # set institution name
    compile_document_button = tk.Button(root, text='Compile NACE Document', command=compile_document)

    root.mainloop()
